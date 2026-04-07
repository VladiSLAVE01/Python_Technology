import csv
from collections import defaultdict
from datetime import datetime
import openpyxl

def analyze_data():
    # Словари для хранения данных
    salary_by_year = defaultdict(list)
    count_by_year = defaultdict(int)
    salary_by_city = defaultdict(list)
    count_by_city = defaultdict(int)
    total_vacancies = 0

    # Чтение CSV файла
    with open('vacancies.csv', 'r', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row in reader:
            if len(row) < 6:
                continue

            # Парсим данные
            vacancy_name = row[0]
            salary_from = float(row[1])
            salary_to = float(row[2])
            currency = row[3]
            city = row[4]
            date_str = row[5]

            # Извлекаем год из даты
            date = datetime.strptime(date_str, '%Y-%m-%dT%H:%M:%S%z')
            year = date.year

            # Пропускаем данные до 2007 года
            if year < 2007 or year > 2022:
                continue

            # Рассчитываем среднюю зарплату (только для RUR)
            if currency == 'RUR':
                avg_salary = (salary_from + salary_to) / 2
                salary_by_year[year].append(avg_salary)
                salary_by_city[city].append(avg_salary)

            # Считаем количество вакансий
            count_by_year[year] += 1
            count_by_city[city] += 1
            total_vacancies += 1

    def process_year_statistics():
        # Вычисляем средние зарплаты по годам
        average_salaries = {}
        for year, salaries in salary_by_year.items():
            average_salaries[year] = sum(salaries) / len(salaries)

        # Сортируем по годам
        years_sorted = sorted(average_salaries.keys())
        counts_sorted = [count_by_year[year] for year in years_sorted]
        salaries_sorted = [average_salaries[year] for year in years_sorted]
        return years_sorted, counts_sorted, salaries_sorted

    def process_city_statistics():
        # Вычисляем средние зарплаты по городам и долю вакансий
        average_salaries_city = {}
        vacancy_shares = {}

        for city, salaries in salary_by_city.items():
            average_salaries_city[city] = sum(salaries) / len(salaries)

        for city, count in count_by_city.items():
            share = (count / total_vacancies) * 100
            vacancy_shares[city] = round(share, 2)

        # Фильтруем города с долей вакансий > 1%
        filtered_cities = {city: share for city, share in vacancy_shares.items() if share > 1.0}

        # Сортируем зарплаты по убыванию, при равенстве по алфавиту
        sorted_salaries_city = sorted(
            [(city, avg_salary) for city, avg_salary in average_salaries_city.items() if city in filtered_cities],
            key=lambda x: (-x[1], x[0])
        )[:10]

        # Сортируем доли вакансий по убыванию, при равенстве по алфавиту
        sorted_shares = sorted(
            [(city, share) for city, share in vacancy_shares.items() if city in filtered_cities],
            key=lambda x: (-x[1], x[0])
        )[:10]

        return sorted_salaries_city, sorted_shares

    def create_excel_file(years_data, cities_data):
        years_sorted, counts_sorted, salaries_sorted = years_data
        sorted_salaries_city, sorted_shares = cities_data

        # Функция для округления зарплат до целого числа
        def round_salary(salary):
            return round(salary)

        # Стили для границ и шрифтов
        from openpyxl.styles import Border, Side, Font
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        header_font = Font(bold=True)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # 1 СТРАНИЦА
        ws_years = wb.create_sheet("Статистика по годам")

        # Заголовки таблицы
        headers_years = ['Год', 'Средняя зарплата', 'Количество вакансий']
        for col, header in enumerate(headers_years, 1):
            cell = ws_years.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.border = thin_border

        # Заполняем данные с границами
        for i, year in enumerate(years_sorted, 2):
            for col in range(1, 4):
                if col == 1:
                    value = year
                elif col == 2:
                    value = round_salary(salaries_sorted[i - 2])
                else:
                    value = counts_sorted[i - 2]
                cell = ws_years.cell(row=i, column=col, value=value)
                cell.border = thin_border

        # Настраиваем ширину столбцов для вкладки по годам
        ws_years.column_dimensions['A'].width = 8
        ws_years.column_dimensions['B'].width = 20
        ws_years.column_dimensions['C'].width = 20

        # 2 СТРАНИЦА
        ws_cities = wb.create_sheet("Статистика по городам")

        # Таблица 1: Уровень зарплат по городам
        headers_salaries = ['Город', 'Уровень зарплат']
        for col, header in enumerate(headers_salaries, 1):
            cell = ws_cities.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.border = thin_border

        # Заполняем данные по зарплатам городов с границами
        for i, (city, salary) in enumerate(sorted_salaries_city, 2):
            for col in range(1, 3):
                if col == 1:
                    value = city
                else:
                    value = round_salary(salary)
                cell = ws_cities.cell(row=i, column=col, value=value)
                cell.border = thin_border

        # Таблица 2: Доля вакансий по городам
        headers_shares = ['Город', 'Доля вакансий, %']
        for col, header in enumerate(headers_shares, 1):
            cell = ws_cities.cell(row=1, column=col + 3, value=header)
            cell.font = header_font
            cell.border = thin_border

        # Заполняем данные по долям вакансий с границами
        for i, (city, share) in enumerate(sorted_shares, 2):
            for col in range(4, 6):
                if col == 4:
                    value = city
                else:
                    value = share
                cell = ws_cities.cell(row=i, column=col, value=value)
                cell.border = thin_border

        # Настраиваем ширину столбцов для вкладки по городам
        ws_cities.column_dimensions['A'].width = 25
        ws_cities.column_dimensions['B'].width = 18
        ws_cities.column_dimensions['C'].width = 5
        ws_cities.column_dimensions['D'].width = 25
        ws_cities.column_dimensions['E'].width = 18

        wb.save('student_works/report.xlsx')

    years_data = process_year_statistics()
    cities_data = process_city_statistics()
    create_excel_file(years_data, cities_data)

analyze_data()
#Сортировку по зарплете с лямбой писал не сам, вроде понимаю как они работают, но самому сотавить является проблемой